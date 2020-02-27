# importing openpyxl module 
import openpyxl as xl;
import re

# opening the source excel file 
filename ="C:\\Users\\abharti\\Desktop\\Phase4-Execution\\Execution_12_Nov_2019\\output_report.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]
ws2 = wb1.worksheets[1]
# intilize a null list
unique_list = []
# function to get unique values
def unique(list1):

    # traverse for all elements
    for x in list1:
        # check if exists in unique_list or not
        if x not in unique_list:
            unique_list.append(x)
            # print list
    for x in unique_list:
        print x,

# calculate total number of rows and
# columns in source excel file
mr = ws1.max_row
mc = ws1.max_column
print mr
print mc
test_script_set = []
Master_test_result = {}

# copying the cell values from source
# excel file to destination excel file
for i in range (1, mr + 1):
        testScriptName = ws1.cell(row = i, column = 3).value
        status = ws1.cell(row = i, column = 4).value
        errorText = ws1.cell(row = i, column = 25).value
        x= re.sub(r"_[a-zA-Z0-9]+$", "", testScriptName.strip(" "))
        #Master_test_result.update({testScriptName: status.value+"="+errorText.value})
        Master_test_result[testScriptName.strip(" ")] = status
        test_script_set.append(x)

colNum = 2
test_script_unique = []
#unique(test_script_set)
uniqueSet = set(test_script_set)
#unique_list = list(dict.fromkeys(test_script_set))
Final_test_result = {}
final_list = []
for x in uniqueSet:
    print x
    ws2.cell(row = colNum, column = 2).value = x
    deviceList = ["TSS7", "TSS10", "TSW560", "TSW560P","TSW760","TSW1060"];
    j = 0
    while(j<len(deviceList)):
     keyName = x+"_"+deviceList[j]
     print "keyname : "+keyName
     status = Master_test_result.get(keyName)
     print status
     if status == None:
     #    ws2.cell(row = colNum, column = 3).value = "NR"
       if x not in Final_test_result:
        Final_test_result[x] = "NR"
       else:
        Final_test_result[x] = Final_test_result.get(x)+":NR"
     else:
     #   ws2.cell(row = colNum, column = 3).value = status
       # Final_test_result[x.strip(" ")] = Final_test_result.get(x)+status
        if x not in Final_test_result:
            Final_test_result[x] = status
        else:
            Final_test_result[x] = Final_test_result.get(x)+":"+status
     j=j+1
     colNum=colNum+1

#Write pass/Fail status in excel worksheet
colNum =2
ws2.cell(row = 1, column = 2).value = "ScriptName"
ws2.cell(row = 1, column = 3).value = "TSS7"
ws2.cell(row = 1, column = 5).value = "TSS10"
ws2.cell(row = 1, column = 7).value = "TSW560"
ws2.cell(row = 1, column = 9).value = "TSW560P"
ws2.cell(row = 1, column = 11).value = "TSW760"
ws2.cell(row = 1, column = 13).value = "TSW1060"
for x in Final_test_result:
    ws2.cell(row = colNum, column = 2).value = x
    ws2.cell(row = colNum, column = 3).value = Final_test_result.get(x).split(":")[0]
    ws2.cell(row = colNum, column = 5).value = Final_test_result.get(x).split(":")[1]
    ws2.cell(row = colNum, column = 7).value = Final_test_result.get(x).split(":")[2]
    ws2.cell(row = colNum, column = 9).value = Final_test_result.get(x).split(":")[3]
    ws2.cell(row = colNum, column = 11).value = Final_test_result.get(x).split(":")[4]
    ws2.cell(row = colNum, column = 13).value = Final_test_result.get(x).split(":")[5]
    colNum=colNum+1
wb1.save(str(filename))

print "--------------------==========================="
print Master_test_result